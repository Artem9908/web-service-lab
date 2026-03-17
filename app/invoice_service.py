from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


HEADER_ORDER_NUMBER = "№ Заказа"
TOTAL_LABEL = "ИТОГО, руб. без учета НДС (20%)"
VAT_LABEL = "НДС (20%)"
TOTAL_WITH_VAT_LABEL = "ИТОГО, руб. с НДС (20%)"


def _parse_decimal(value: str | int | float | Decimal) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    normalized = str(value).replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        raise ValueError(f"Не удалось преобразовать число: {value}") from None


def _replace_header_meta(source: str, date_str: str, invoice_number: int) -> str:
    date_pattern = r"(Дата счета:\s*)(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})"
    invoice_pattern = r"(Номер счета:\s*)(\d+)"

    updated = re.sub(date_pattern, rf"\g<1>{date_str}", source)
    updated = re.sub(invoice_pattern, rf"\g<1>{invoice_number}", updated)
    return updated


def _find_row_by_text(sheet, text: str, col_start: int = 1, col_end: int = 7) -> int:
    for row in range(1, sheet.max_row + 1):
        for col in range(col_start, col_end + 1):
            value = sheet.cell(row=row, column=col).value
            if isinstance(value, str) and text in value:
                return row
    raise ValueError(f"Не удалось найти строку с текстом: {text}")


def _copy_row_style(sheet, src_row: int, dst_row: int, col_start: int = 1, col_end: int = 7) -> None:
    for col in range(col_start, col_end + 1):
        src = sheet.cell(row=src_row, column=col)
        dst = sheet.cell(row=dst_row, column=col)
        dst._style = copy(src._style)


def _rows_to_decimal_total(rows: Iterable[Sequence[str]]) -> Decimal:
    total = Decimal("0")
    for row in rows:
        total += _parse_decimal(row[5])
    return total


def _clear_dynamic_merges(sheet, start_row: int) -> None:
    # В динамической части шаблона удаляем все merge-диапазоны,
    # иначе при 10+ строках старые объединения ломают таблицу.
    ranges_to_remove = []
    for merged_range in sheet.merged_cells.ranges:
        _, min_row, _, max_row = merged_range.bounds
        if max_row >= start_row:
            ranges_to_remove.append(str(merged_range))
    for merged in ranges_to_remove:
        sheet.unmerge_cells(merged)


def _reset_total_merges(sheet, total_row: int) -> None:
    # После delete/insert openpyxl может оставить некорректные merged-диапазоны.
    # Явно пересобираем объединения для строк итогов A:F.
    ranges_to_remove = []
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col == 1 and max_col == 6 and max_row >= total_row and min_row <= total_row + 2:
            ranges_to_remove.append(str(merged_range))
    for merged in ranges_to_remove:
        sheet.unmerge_cells(merged)

    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    sheet.merge_cells(start_row=total_row + 1, start_column=1, end_row=total_row + 1, end_column=6)
    sheet.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=6)


def generate_invoice(
    template_path: Path,
    output_dir: Path,
    date_iso: str,
    invoice_number: int,
    period: str,
    data_rows: list[list[str]],
) -> Path:
    if not data_rows:
        raise ValueError("Поле data не должно быть пустым")

    wb = load_workbook(template_path)
    ws = wb.active

    date_as_str = datetime.strptime(date_iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    ws["G1"] = _replace_header_meta(str(ws["G1"].value), date_as_str, invoice_number)
    ws["G2"] = f"Период: {period}"

    header_row = _find_row_by_text(ws, HEADER_ORDER_NUMBER)
    detail_start = header_row + 1
    total_row = _find_row_by_text(ws, TOTAL_LABEL)
    _clear_dynamic_merges(ws, detail_start)

    # Сохраняем стиль первой строки таблицы, чтобы применить его к новым строкам.
    style_anchor_row = detail_start
    style_cache = [copy(ws.cell(style_anchor_row, col)._style) for col in range(1, 8)]
    total_style_cache = [copy(ws.cell(total_row, col)._style) for col in range(1, 8)]
    vat_style_cache = [copy(ws.cell(total_row + 1, col)._style) for col in range(1, 8)]
    total_with_vat_style_cache = [copy(ws.cell(total_row + 2, col)._style) for col in range(1, 8)]

    existing_count = max(total_row - detail_start, 0)
    if existing_count:
        ws.delete_rows(detail_start, existing_count)

    ws.insert_rows(detail_start, amount=len(data_rows))

    for index, item in enumerate(data_rows, start=1):
        if len(item) != 6:
            raise ValueError("Каждая строка в data должна содержать 6 элементов")
        row = detail_start + index - 1
        ws.cell(row=row, column=1).value = index
        ws.cell(row=row, column=2).value = item[0]
        ws.cell(row=row, column=3).value = item[1]
        ws.cell(row=row, column=4).value = item[2]
        ws.cell(row=row, column=5).value = item[3]
        ws.cell(row=row, column=6).value = float(_parse_decimal(item[4]))
        ws.cell(row=row, column=7).value = float(_parse_decimal(item[5]))
        for col in range(1, 8):
            ws.cell(row=row, column=col)._style = copy(style_cache[col - 1])

    total_row = detail_start + len(data_rows)
    for col in range(1, 8):
        ws.cell(total_row, col)._style = copy(total_style_cache[col - 1])
        ws.cell(total_row + 1, col)._style = copy(vat_style_cache[col - 1])
        ws.cell(total_row + 2, col)._style = copy(total_with_vat_style_cache[col - 1])
    _reset_total_merges(ws, total_row)

    ws.cell(total_row, 1).value = TOTAL_LABEL
    ws.cell(total_row + 1, 1).value = VAT_LABEL
    ws.cell(total_row + 2, 1).value = TOTAL_WITH_VAT_LABEL

    subtotal = _rows_to_decimal_total(data_rows)
    vat = (subtotal * Decimal("0.2")).quantize(Decimal("0.01"))
    total = (subtotal + vat).quantize(Decimal("0.01"))
    ws.cell(total_row, 7).value = float(subtotal)
    ws.cell(total_row + 1, 7).value = float(vat)
    ws.cell(total_row + 2, 7).value = float(total)

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"invoice_{invoice_number}_{ts}.xlsx"
    wb.save(output_path)
    return output_path
