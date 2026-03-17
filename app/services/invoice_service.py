from __future__ import annotations

import re
from collections.abc import Iterable, Sequence
from copy import copy
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

HEADER_ORDER_NUMBER = "№ Заказа"
TOTAL_LABEL = "ИТОГО, руб. без учета НДС (20%)"
VAT_LABEL = "НДС (20%)"
TOTAL_WITH_VAT_LABEL = "ИТОГО, руб. с НДС (20%)"
PLACEHOLDER_PATTERN = re.compile(r"^\{\{([a-zA-Z0-9_]+)\}\}$")


class InvoiceServiceError(ValueError):
    """Business-level error for invoice generation."""


def _parse_decimal(value: str | int | float | Decimal) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    normalized = str(value).replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise InvoiceServiceError(f"Не удалось преобразовать число: {value}") from exc


def _replace_header_meta(source: str, date_str: str, invoice_number: int) -> str:
    date_pattern = r"(Дата счета:\s*)(\d{2}\.\d{2}\.\d{4}|\d{4}-\d{2}-\d{2})"
    invoice_pattern = r"(Номер счета:\s*)(\d+)"

    updated = re.sub(date_pattern, rf"\g<1>{date_str}", source)
    updated = re.sub(invoice_pattern, rf"\g<1>{invoice_number}", updated)
    return updated


def _replace_text_placeholders(sheet: Worksheet, mapping: dict[str, str]) -> int:
    replacements = 0
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if not isinstance(cell.value, str) or "{{" not in cell.value:
                continue
            original = cell.value
            updated = original
            for key, value in mapping.items():
                updated = updated.replace(f"{{{{{key}}}}}", value)
            if updated != original:
                cell.value = updated
                replacements += 1
    return replacements


def _extract_row_placeholders(
    sheet: Worksheet, row: int, col_start: int = 1, col_end: int = 7
) -> dict[int, str]:
    placeholders: dict[int, str] = {}
    for col in range(col_start, col_end + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if not isinstance(cell_value, str):
            continue
        match = PLACEHOLDER_PATTERN.match(cell_value.strip())
        if match:
            placeholders[col] = match.group(1)
    return placeholders


def _build_row_context(index: int, item: Sequence[str]) -> dict[str, str | int | float]:
    order_id = item[0]
    service_id = item[1]
    device_name = item[2]
    period = item[3]
    sum_value = float(_parse_decimal(item[4]))
    total_sum_value = float(_parse_decimal(item[5]))

    return {
        "index": index,
        "order_id": order_id,
        "service_id": service_id,
        "device_name": device_name,
        "period": period,
        "sum": sum_value,
        "total_sum": total_sum_value,
        # Legacy aliases used in some Excel templates.
        "data0": order_id,
        "data1": service_id,
        "data2": index,
        "data3": device_name,
        "data5": period,
        "data7": sum_value,
        "data10": total_sum_value,
    }


def _resolve_row_value(
    col: int,
    row_placeholders: dict[int, str],
    row_context: dict[str, str | int | float],
    item: Sequence[str],
    index: int,
) -> str | int | float:
    if col in row_placeholders:
        placeholder_key = row_placeholders[col]
        if placeholder_key not in row_context:
            raise InvoiceServiceError(f"Неизвестный плейсхолдер в строке данных: {placeholder_key}")
        return row_context[placeholder_key]

    if col == 1:
        return index
    if col == 2:
        return item[0]
    if col == 3:
        return item[1]
    if col == 4:
        return item[2]
    if col == 5:
        return item[3]
    if col == 6:
        return float(_parse_decimal(item[4]))
    if col == 7:
        return float(_parse_decimal(item[5]))
    raise InvoiceServiceError(f"Неподдерживаемая колонка таблицы: {col}")


def _find_row_by_text(sheet: Worksheet, text: str, col_start: int = 1, col_end: int = 7) -> int:
    for row in range(1, sheet.max_row + 1):
        for col in range(col_start, col_end + 1):
            value = sheet.cell(row=row, column=col).value
            if isinstance(value, str) and text in value:
                return row
    raise InvoiceServiceError(f"Не удалось найти строку с текстом: {text}")


def _rows_to_decimal_total(rows: Iterable[Sequence[str]]) -> Decimal:
    total = Decimal("0")
    for row in rows:
        total += _parse_decimal(row[5])
    return total


def _clear_dynamic_merges(sheet: Worksheet, start_row: int) -> None:
    ranges_to_remove: list[str] = []
    for merged_range in sheet.merged_cells.ranges:
        _, _, _, max_row = merged_range.bounds
        if max_row >= start_row:
            ranges_to_remove.append(str(merged_range))
    for merged in ranges_to_remove:
        sheet.unmerge_cells(merged)


def _reset_total_merges(sheet: Worksheet, total_row: int) -> None:
    ranges_to_remove: list[str] = []
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col == 1 and max_col == 6 and max_row >= total_row and min_row <= total_row + 2:
            ranges_to_remove.append(str(merged_range))
    for merged in ranges_to_remove:
        sheet.unmerge_cells(merged)

    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    sheet.merge_cells(start_row=total_row + 1, start_column=1, end_row=total_row + 1, end_column=6)
    sheet.merge_cells(start_row=total_row + 2, start_column=1, end_row=total_row + 2, end_column=6)


def generate_invoice(
    template_path: Path,
    output_dir: Path,
    date_iso: str,
    invoice_number: int,
    period: str,
    data_rows: list[list[str]],
) -> Path:
    if not template_path.exists():
        raise InvoiceServiceError(f"Шаблон не найден: {template_path}")
    if not data_rows:
        raise InvoiceServiceError("Поле data не должно быть пустым")

    try:
        date_as_str = datetime.strptime(date_iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError as exc:
        raise InvoiceServiceError("Неверный формат даты, ожидается YYYY-MM-DD") from exc

    wb = load_workbook(template_path)
    ws = wb.active

    placeholder_replacements = _replace_text_placeholders(
        ws,
        {
            "invoice_date": date_as_str,
            "invoice_number": str(invoice_number),
            "period": period,
        },
    )

    # Backward compatibility для старых шаблонов без {{placeholders}}.
    if placeholder_replacements == 0:
        ws["G1"] = _replace_header_meta(str(ws["G1"].value), date_as_str, invoice_number)
        ws["G2"] = f"Период: {period}"

    header_row = _find_row_by_text(ws, HEADER_ORDER_NUMBER)
    detail_start = header_row + 1
    total_row = _find_row_by_text(ws, TOTAL_LABEL)
    _clear_dynamic_merges(ws, detail_start)
    row_placeholders = _extract_row_placeholders(ws, detail_start)

    style_anchor_row = detail_start
    style_cache = [copy(ws.cell(style_anchor_row, col)._style) for col in range(1, 8)]
    total_style_cache = [copy(ws.cell(total_row, col)._style) for col in range(1, 8)]
    vat_style_cache = [copy(ws.cell(total_row + 1, col)._style) for col in range(1, 8)]
    total_with_vat_style_cache = [copy(ws.cell(total_row + 2, col)._style) for col in range(1, 8)]

    existing_count = max(total_row - detail_start, 0)
    if existing_count:
        ws.delete_rows(detail_start, existing_count)

    ws.insert_rows(detail_start, amount=len(data_rows))

    for index, item in enumerate(data_rows, start=1):
        if len(item) != 6:
            raise InvoiceServiceError("Каждая строка в data должна содержать 6 элементов")
        row = detail_start + index - 1
        row_context = _build_row_context(index, item)
        for col in range(1, 8):
            ws.cell(row=row, column=col).value = _resolve_row_value(
                col=col,
                row_placeholders=row_placeholders,
                row_context=row_context,
                item=item,
                index=index,
            )
        for col in range(1, 8):
            ws.cell(row=row, column=col)._style = copy(style_cache[col - 1])

    total_row = detail_start + len(data_rows)
    for col in range(1, 8):
        ws.cell(total_row, col)._style = copy(total_style_cache[col - 1])
        ws.cell(total_row + 1, col)._style = copy(vat_style_cache[col - 1])
        ws.cell(total_row + 2, col)._style = copy(total_with_vat_style_cache[col - 1])
    _reset_total_merges(ws, total_row)

    ws.cell(total_row, 1).value = TOTAL_LABEL
    ws.cell(total_row + 1, 1).value = VAT_LABEL
    ws.cell(total_row + 2, 1).value = TOTAL_WITH_VAT_LABEL

    subtotal = _rows_to_decimal_total(data_rows)
    vat = (subtotal * Decimal("0.2")).quantize(Decimal("0.01"))
    total = (subtotal + vat).quantize(Decimal("0.01"))
    ws.cell(total_row, 7).value = float(subtotal)
    ws.cell(total_row + 1, 7).value = float(vat)
    ws.cell(total_row + 2, 7).value = float(total)

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    output_path = output_dir / f"invoice_{invoice_number}_{ts}.xlsx"
    wb.save(output_path)
    return output_path
