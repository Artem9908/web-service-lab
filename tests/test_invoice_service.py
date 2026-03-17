from pathlib import Path

from openpyxl import load_workbook

from app.services.invoice_service import generate_invoice


def test_generate_invoice_with_many_rows(tmp_path: Path) -> None:
    data_rows = [
        [
            f"М-{index}",
            f"700000{index:05d}",
            f"ID {index}",
            "01.10.2025 - 31.10.2025",
            "1000",
            f"{1000 + index},00",
        ]
        for index in range(1, 13)
    ]

    output = generate_invoice(
        template_path=Path("templates/invoice_template.xlsx"),
        output_dir=tmp_path,
        date_iso="2025-10-31",
        invoice_number=1001,
        period="Октябрь 2025",
        data_rows=data_rows,
    )

    assert output.exists()
    workbook = load_workbook(output)
    sheet = workbook.active

    assert "Номер счета: 1001" in str(sheet["G1"].value)
    assert sheet["G2"].value == "Период: Октябрь 2025"
    assert sheet["A19"].value == "ИТОГО, руб. без учета НДС (20%)"
    assert sheet["A20"].value == "НДС (20%)"
    assert sheet["A21"].value == "ИТОГО, руб. с НДС (20%)"
    assert sheet["G19"].value > 0
    assert sheet["G20"].value > 0
    assert sheet["G21"].value > 0


def test_generate_invoice_replaces_template_placeholders(tmp_path: Path) -> None:
    template_copy = tmp_path / "invoice_template_placeholders.xlsx"
    original = load_workbook("templates/invoice_template.xlsx")
    original_sheet = original.active
    original_sheet["G1"] = (
        "Договор № D240352920\n" "Дата счета: {{invoice_date}}\n" "Номер счета: {{invoice_number}}"
    )
    original_sheet["G2"] = "Период: {{period}}"
    original.save(template_copy)

    output = generate_invoice(
        template_path=template_copy,
        output_dir=tmp_path,
        date_iso="2026-02-28",
        invoice_number=1200,
        period="Февраль 2026",
        data_rows=[
            ["М-1", "7001", "ID 1", "01.02.2026 - 28.02.2026", "1000", "1000,00"],
        ],
    )

    workbook = load_workbook(output)
    sheet = workbook.active
    assert "{{invoice_date}}" not in str(sheet["G1"].value)
    assert "{{invoice_number}}" not in str(sheet["G1"].value)
    assert "{{period}}" not in str(sheet["G2"].value)
    assert "Дата счета: 28.02.2026" in str(sheet["G1"].value)
    assert "Номер счета: 1200" in str(sheet["G1"].value)
    assert sheet["G2"].value == "Период: Февраль 2026"
